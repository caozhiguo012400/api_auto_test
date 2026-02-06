# -*- coding:utf-8 -*-
import openpyxl
import pandas as pd
import os
from core.logger import log
from core.exception_handler import DataNotFoundError, ConfigLoadError
from utils.path_util import get_project_path


class ExcelUtil:
    """Excel工具类（封装Excel读取/写入，支持批量读取测试数据、生成用例）"""

    def __init__(self, file_path: str):
        """
        初始化Excel工具
        :param file_path: Excel文件路径（相对路径/绝对路径，相对路径基于项目根目录）
        """
        # 处理文件路径（相对路径转换为绝对路径）
        if not os.path.isabs(file_path):
            self.file_path = os.path.join(get_project_path(), file_path)
        else:
            self.file_path = file_path
        # 校验文件是否存在
        if not os.path.exists(self.file_path):
            raise DataNotFoundError(f"Excel文件不存在：{self.file_path}")
        # 校验文件格式（必须是xlsx格式）
        if not self.file_path.endswith(".xlsx"):
            raise ConfigLoadError(f"Excel文件格式错误：仅支持xlsx格式，当前文件：{self.file_path}")

        self.workbook = None  # Excel工作簿对象
        self.sheet = None  # Excel工作表对象

    def open_excel(self, sheet_name: str = None):
        """
        打开Excel文件，指定工作表（默认打开第一个工作表）
        :param sheet_name: 工作表名称（如"登录用例"）
        """
        try:
            # 打开Excel文件（read_only=False，支持写入）
            self.workbook = openpyxl.load_workbook(self.file_path)
            # 指定工作表
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    raise DataNotFoundError(f"Excel工作表不存在：{sheet_name}，文件：{self.file_path}")
                self.sheet = self.workbook[sheet_name]
            else:
                # 默认打开第一个工作表
                self.sheet = self.workbook.active
            log.info(f"成功打开Excel文件：{self.file_path}，工作表：{self.sheet.title}")
        except Exception as e:
            raise ConfigLoadError(f"打开Excel文件失败：{str(e)}，文件：{self.file_path}")

    def read_excel_to_dict(self, sheet_name: str = None, header_row: int = 1) -> list:
        """
        读取Excel数据，转换为字典列表（最常用，便于pytest参数化）
        :param sheet_name: 工作表名称
        :param header_row: 表头行号（默认第1行，表头为字典的key）
        :return: 字典列表（每个元素对应一行数据，key为表头，value为单元格值）
        """
        self.open_excel(sheet_name)
        try:
            # 获取表头（header_row行的内容）
            headers = []
            for col in range(1, self.sheet.max_column + 1):
                header_value = self.sheet.cell(row=header_row, column=col).value
                if not header_value:
                    raise ConfigLoadError(f"Excel表头异常：第{header_row}行第{col}列表头为空")
                headers.append(str(header_value).strip())

            # 读取表头以下的数据，转换为字典列表
            data_list = []
            # 从表头下一行开始读取（header_row + 1）
            for row in range(header_row + 1, self.sheet.max_row + 1):
                row_data = {}
                for col in range(1, self.sheet.max_column + 1):
                    # 获取单元格值（处理空值为None）
                    cell_value = self.sheet.cell(row=row, column=col).value
                    row_data[headers[col - 1]] = cell_value
                # 跳过空行（所有字段都为空的行）
                if any(row_data.values()):
                    data_list.append(row_data)

            log.info(f"读取Excel数据成功：共{len(data_list)}条数据，工作表：{self.sheet.title}")
            return data_list
        except Exception as e:
            raise DataNotFoundError(f"读取Excel数据失败：{str(e)}，文件：{self.file_path}")
        finally:
            # 关闭Excel工作簿，释放资源
            if self.workbook:
                self.workbook.close()

    def write_excel(self, data_list: list, sheet_name: str = "test_data", header_row: int = 1):
        """
        写入数据到Excel文件（覆盖工作表内容，支持批量写入）
        :param data_list: 待写入的数据（字典列表，所有字典的key需一致）
        :param sheet_name: 工作表名称（不存在则创建）
        :param header_row: 表头行号（默认第1行）
        """
        if not data_list or not isinstance(data_list, list):
            raise DataNotFoundError("写入Excel失败：待写入数据为空或不是列表类型")
        if not all(isinstance(item, dict) for item in data_list):
            raise DataNotFoundError("写入Excel失败：待写入数据列表中的元素必须是字典类型")

        try:
            # 打开Excel文件（若不存在则创建）
            if os.path.exists(self.file_path):
                self.workbook = openpyxl.load_workbook(self.file_path)
            else:
                self.workbook = openpyxl.Workbook()

            # 处理工作表（存在则删除，重新创建，避免内容重复）
            if sheet_name in self.workbook.sheetnames:
                self.workbook.remove(self.workbook[sheet_name])
            self.sheet = self.workbook.create_sheet(title=sheet_name)

            # 获取表头（所有字典的key的并集，保证所有字段都能写入）
            headers = []
            for item in data_list:
                for key in item.keys():
                    if key not in headers:
                        headers.append(str(key).strip())

            # 写入表头
            for col in range(1, len(headers) + 1):
                self.sheet.cell(row=header_row, column=col, value=headers[col - 1])

            # 写入数据（从表头下一行开始）
            for row_idx, row_data in enumerate(data_list, start=header_row + 1):
                for col_idx, header in enumerate(headers, start=1):
                    # 写入对应字段的值（不存在则为空）
                    cell_value = row_data.get(header, None)
                    self.sheet.cell(row=row_idx, column=col_idx, value=cell_value)

            # 保存文件
            self.workbook.save(self.file_path)
            log.info(f"写入Excel数据成功：共写入{len(data_list)}条数据，工作表：{sheet_name}，文件：{self.file_path}")
        except Exception as e:
            raise ConfigLoadError(f"写入Excel数据失败：{str(e)}，文件：{self.file_path}")
        finally:
            if self.workbook:
                self.workbook.close()

    @staticmethod
    def batch_read_excel(folder_path: str = "data/excel_data") -> dict:
        """
        批量读取指定文件夹下所有Excel文件的数据
        :param folder_path: Excel文件所在文件夹路径（默认data/excel_data）
        :return: 字典（key：文件名，value：该文件下所有工作表的数据）
        """
        # 处理文件夹路径
        folder_path = os.path.join(get_project_path(), folder_path)
        if not os.path.exists(folder_path):
            raise DataNotFoundError(f"批量读取Excel失败：文件夹不存在：{folder_path}")

        all_excel_data = {}
        # 遍历文件夹下所有xlsx文件
        for file_name in os.listdir(folder_path):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(folder_path, file_name)
                excel_util = ExcelUtil(file_path)
                # 读取该文件下所有工作表的数据
                workbook = openpyxl.load_workbook(file_path)
                sheet_data = {}
                for sheet_name in workbook.sheetnames:
                    sheet_data[sheet_name] = excel_util.read_excel_to_dict(sheet_name=sheet_name)
                workbook.close()
                # 以文件名作为key，存入总字典（去除后缀）
                file_key = os.path.splitext(file_name)[0]
                all_excel_data[file_key] = sheet_data

        log.info(f"批量读取Excel完成：共读取{len(all_excel_data)}个Excel文件，文件夹：{folder_path}")
        return all_excel_data


# 导出Excel工具实例（全局复用）
excel_util = ExcelUtil(file_path="data/excel_data/user_api_data.xlsx")